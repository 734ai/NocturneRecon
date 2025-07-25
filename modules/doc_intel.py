#!/usr/bin/env python3
"""
NocturneRecon Document Intelligence Module
Author: Muzan Sano
Version: 2.0.0-dev
License: MIT

This module provides document intelligence capabilities including:
- PDF metadata extraction
- Office document analysis
- Public file discovery
- Document leak detection
"""

import os
import sys
import time
import requests
import threading
import re
from urllib.parse import urljoin, urlparse, quote
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.utils import print_info, print_success, print_warning, print_error, save_to_json, save_to_csv, save_to_txt

try:
    import PyPDF2
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print_warning("PyPDF2 not available - PDF analysis will be limited")

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print_warning("openpyxl not available - Excel analysis will be limited")

try:
    from docx import Document as DocxDocument
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False
    print_warning("python-docx not available - Word document analysis will be limited")


class DocumentIntelligenceGatherer:
    """
    Document Intelligence Gathering class for discovering and analyzing public documents
    """
    
    def __init__(self, target, config=None):
        """
        Initialize the Document Intelligence Gatherer
        
        Args:
            target (str): Target domain or organization to search for
            config (dict): Configuration dictionary
        """
        self.target = target
        self.config = config or {}
        self.results = {
            'target': target,
            'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
            'discovered_documents': [],
            'pdf_analysis': [],
            'office_analysis': [],
            'document_metadata': [],
            'leaked_documents': [],
            'sensitive_content': []
        }
        
        # Document file extensions to search for
        self.document_extensions = {
            'pdf': ['pdf'],
            'office': ['doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx'],
            'text': ['txt', 'rtf'],
            'other': ['zip', 'rar', '7z', 'tar.gz']
        }
        
        # Search engines and sources
        self.search_sources = [
            'google.com',
            'bing.com',
            'duckduckgo.com',
            'archive.org'
        ]
        
        # Document discovery patterns
        self.search_patterns = {
            'general_docs': [
                f'site:{self.target} filetype:pdf',
                f'site:{self.target} filetype:doc',
                f'site:{self.target} filetype:docx',
                f'site:{self.target} filetype:xls',
                f'site:{self.target} filetype:xlsx',
                f'site:{self.target} filetype:ppt',
                f'site:{self.target} filetype:pptx'
            ],
            'sensitive_docs': [
                f'{self.target} "confidential" filetype:pdf',
                f'{self.target} "internal" filetype:doc',
                f'{self.target} "private" filetype:pdf',
                f'{self.target} "password" filetype:pdf',
                f'{self.target} "employee" filetype:pdf',
                f'{self.target} "financial" filetype:xlsx'
            ],
            'technical_docs': [
                f'{self.target} "manual" filetype:pdf',
                f'{self.target} "guide" filetype:pdf',
                f'{self.target} "documentation" filetype:pdf',
                f'{self.target} "api" filetype:pdf',
                f'{self.target} "technical" filetype:pdf'
            ]
        }
        
        # Metadata fields to extract
        self.metadata_fields = [
            'title', 'author', 'creator', 'producer', 'subject',
            'keywords', 'creation_date', 'modification_date',
            'company', 'manager', 'category', 'comments'
        ]
        
        # Sensitive content patterns
        self.sensitive_patterns = {
            'emails': r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
            'phone_numbers': r'(\+\d{1,3}[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}',
            'ip_addresses': r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b',
            'urls': r'https?://[^\s<>"\']+',
            'social_security': r'\b\d{3}-\d{2}-\d{4}\b',
            'credit_cards': r'\b(?:4[0-9]{12}(?:[0-9]{3})?|5[1-5][0-9]{14}|3[47][0-9]{13})\b',
            'api_keys': r'(?i)(api[_-]?key|access[_-]?token)\s*[:=]\s*["\']?([a-zA-Z0-9_-]+)["\']?',
            'passwords': r'(?i)(password|pass|pwd)\s*[:=]\s*["\']?([^"\'\s]+)["\']?'
        }
        
        self.threads = self.config.get('threads', 3)
        self.delay = self.config.get('delay', 2.0)
        self.timeout = self.config.get('timeout', 20)
        self.max_file_size = self.config.get('max_file_size', 10 * 1024 * 1024)  # 10MB
        
        # Setup session with retry strategy
        self.session = requests.Session()
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)
        
        # User agents for stealth
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        ]

    def discover_public_documents(self):
        """
        Discover public documents across search engines
        """
        print_info(f"Discovering public documents for '{self.target}'...")
        
        all_patterns = []
        for category, patterns in self.search_patterns.items():
            all_patterns.extend(patterns)
        
        for pattern in all_patterns:
            for source in self.search_sources:
                try:
                    print_info(f"Searching {source} for: {pattern}")
                    
                    if source == 'google.com':
                        search_url = f"https://www.google.com/search?q={quote(pattern)}"
                    elif source == 'bing.com':
                        search_url = f"https://www.bing.com/search?q={quote(pattern)}"
                    elif source == 'duckduckgo.com':
                        search_url = f"https://duckduckgo.com/?q={quote(pattern)}"
                    elif source == 'archive.org':
                        search_url = f"https://archive.org/search.php?query={quote(pattern)}"
                    else:
                        continue
                    
                    headers = {
                        'User-Agent': self.user_agents[0],
                        'Accept': 'text/html,application/xhtml+xml'
                    }
                    
                    response = self.session.get(
                        search_url,
                        headers=headers,
                        timeout=self.timeout
                    )
                    
                    if response.status_code == 200:
                        documents = self._extract_document_links(response.text, source, pattern)
                        
                        for doc in documents:
                            # Analyze document
                            doc_analysis = self._analyze_document_url(doc)
                            
                            document_result = {
                                'url': doc.get('url', ''),
                                'title': doc.get('title', ''),
                                'file_type': doc_analysis.get('file_type', ''),
                                'file_size': doc_analysis.get('file_size', ''),
                                'source': source,
                                'search_pattern': pattern,
                                'is_sensitive': doc_analysis.get('is_sensitive', False),
                                'discovered_at': time.strftime('%Y-%m-%d %H:%M:%S')
                            }
                            
                            self.results['discovered_documents'].append(document_result)
                            
                            if doc_analysis.get('is_sensitive'):
                                print_warning(f"Sensitive document found: {doc.get('title', 'Unknown')}")
                            else:
                                print_success(f"Document found: {doc.get('title', 'Unknown')}")
                    
                    time.sleep(self.delay)
                    
                except Exception as e:
                    print_warning(f"Error searching {source} for '{pattern}': {str(e)}")
                    continue

    def analyze_discovered_documents(self):
        """
        Download and analyze discovered documents
        """
        print_info("Analyzing discovered documents...")
        
        for doc in self.results['discovered_documents']:
            try:
                doc_url = doc.get('url', '')
                file_type = doc.get('file_type', '').lower()
                
                if not doc_url:
                    continue
                
                print_info(f"Analyzing document: {doc.get('title', 'Unknown')}")
                
                # Download document
                document_content = self._download_document(doc_url)
                
                if document_content:
                    # Analyze based on file type
                    if file_type == 'pdf' and PDF_SUPPORT:
                        analysis = self._analyze_pdf_document(document_content, doc)
                        if analysis:
                            self.results['pdf_analysis'].append(analysis)
                    
                    elif file_type in ['doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx']:
                        analysis = self._analyze_office_document(document_content, doc)
                        if analysis:
                            self.results['office_analysis'].append(analysis)
                    
                    # Extract metadata
                    metadata = self._extract_document_metadata(document_content, file_type, doc)
                    if metadata:
                        self.results['document_metadata'].append(metadata)
                    
                    # Check for sensitive content
                    sensitive_content = self._scan_for_sensitive_content(document_content, doc)
                    if sensitive_content:
                        self.results['sensitive_content'].extend(sensitive_content)
                
                time.sleep(self.delay)
                
            except Exception as e:
                print_warning(f"Error analyzing document {doc.get('url', '')}: {str(e)}")
                continue

    def scan_for_leaked_documents(self):
        """
        Scan for leaked documents on file sharing sites
        """
        print_info(f"Scanning for leaked documents related to '{self.target}'...")
        
        leak_sources = [
            'scribd.com',
            'slideshare.net',
            'documentcloud.org',
            'issuu.com',
            'mega.nz'
        ]
        
        leak_patterns = [
            f'{self.target} confidential',
            f'{self.target} internal',
            f'{self.target} leaked',
            f'{self.target} dump',
            f'{self.target} private'
        ]
        
        for source in leak_sources:
            for pattern in leak_patterns:
                try:
                    print_info(f"Checking {source} for: {pattern}")
                    
                    search_query = f'site:{source} "{pattern}"'
                    search_url = f"https://www.google.com/search?q={quote(search_query)}"
                    
                    headers = {
                        'User-Agent': self.user_agents[0],
                        'Accept': 'text/html,application/xhtml+xml'
                    }
                    
                    response = self.session.get(
                        search_url,
                        headers=headers,
                        timeout=self.timeout
                    )
                    
                    if response.status_code == 200:
                        leaked_docs = self._extract_leaked_documents(response.text, source, pattern)
                        
                        for doc in leaked_docs:
                            leak_result = {
                                'title': doc.get('title', ''),
                                'url': doc.get('url', ''),
                                'source': source,
                                'leak_type': 'document_sharing',
                                'risk_level': self._assess_leak_risk(doc),
                                'search_pattern': pattern,
                                'discovered_at': time.strftime('%Y-%m-%d %H:%M:%S')
                            }
                            
                            self.results['leaked_documents'].append(leak_result)
                            print_warning(f"Leaked document found: {doc.get('title', 'Unknown')}")
                    
                    time.sleep(self.delay)
                    
                except Exception as e:
                    print_warning(f"Error checking {source} for '{pattern}': {str(e)}")
                    continue

    def _extract_document_links(self, html_content, source, pattern):
        """
        Extract document links from search results
        
        Args:
            html_content (str): HTML content from search page
            source (str): Search engine source
            pattern (str): Search pattern used
            
        Returns:
            list: List of document dictionaries
        """
        documents = []
        
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Extract links based on source
            if 'google.com' in source:
                results = soup.find_all('div', class_='g')
            elif 'bing.com' in source:
                results = soup.find_all('li', class_='b_algo')
            elif 'duckduckgo.com' in source:
                results = soup.find_all('div', class_='result')
            elif 'archive.org' in source:
                results = soup.find_all('div', class_='item-ttl')
            else:
                results = soup.find_all(['div', 'li'], class_=re.compile(r'result|item'))
            
            for result in results:
                title_elem = result.find(['h3', 'h2', 'a'])
                title = title_elem.get_text().strip() if title_elem else 'No title'
                
                link_elem = result.find('a', href=True)
                url = link_elem['href'] if link_elem else ''
                
                # Filter for document URLs
                if url and any(ext in url.lower() for ext_list in self.document_extensions.values() for ext in ext_list):
                    documents.append({
                        'title': title,
                        'url': url
                    })
            
        except Exception as e:
            print_warning(f"Error extracting document links from {source}: {str(e)}")
        
        return documents

    def _analyze_document_url(self, doc):
        """
        Analyze document URL for basic information
        
        Args:
            doc (dict): Document information
            
        Returns:
            dict: Analysis results
        """
        url = doc.get('url', '')
        title = doc.get('title', '').lower()
        
        # Determine file type
        file_type = 'unknown'
        for category, extensions in self.document_extensions.items():
            for ext in extensions:
                if url.lower().endswith(f'.{ext}'):
                    file_type = ext
                    break
        
        # Check if potentially sensitive
        sensitive_keywords = ['confidential', 'internal', 'private', 'secret', 'restricted']
        is_sensitive = any(keyword in title for keyword in sensitive_keywords)
        
        return {
            'file_type': file_type,
            'is_sensitive': is_sensitive,
            'file_size': 'unknown'
        }

    def _download_document(self, url):
        """
        Download document content
        
        Args:
            url (str): Document URL
            
        Returns:
            bytes: Document content or None
        """
        try:
            headers = {
                'User-Agent': self.user_agents[0],
                'Accept': 'application/pdf,application/msword,application/vnd.openxmlformats-officedocument.*,*/*'
            }
            
            response = self.session.get(
                url,
                headers=headers,
                timeout=self.timeout,
                stream=True
            )
            
            if response.status_code == 200:
                # Check file size
                content_length = response.headers.get('content-length')
                if content_length and int(content_length) > self.max_file_size:
                    print_warning(f"Document too large ({content_length} bytes): {url}")
                    return None
                
                # Download content
                content = b''
                for chunk in response.iter_content(chunk_size=8192):
                    content += chunk
                    if len(content) > self.max_file_size:
                        print_warning(f"Document too large during download: {url}")
                        return None
                
                return content
            
        except Exception as e:
            print_warning(f"Error downloading document {url}: {str(e)}")
        
        return None

    def _analyze_pdf_document(self, content, doc_info):
        """
        Analyze PDF document content and metadata
        
        Args:
            content (bytes): PDF content
            doc_info (dict): Document information
            
        Returns:
            dict: PDF analysis results
        """
        if not PDF_SUPPORT:
            return None
        
        try:
            import io
            pdf_file = io.BytesIO(content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            # Extract metadata
            metadata = pdf_reader.metadata if pdf_reader.metadata else {}
            
            # Extract text content
            text_content = ''
            for page in pdf_reader.pages:
                try:
                    text_content += page.extract_text() + '\n'
                except:
                    continue
            
            # Analyze content
            sensitive_data = self._analyze_text_for_sensitive_data(text_content)
            
            analysis = {
                'url': doc_info.get('url', ''),
                'title': doc_info.get('title', ''),
                'page_count': len(pdf_reader.pages),
                'metadata': {
                    'title': str(metadata.get('/Title', '')),
                    'author': str(metadata.get('/Author', '')),
                    'creator': str(metadata.get('/Creator', '')),
                    'producer': str(metadata.get('/Producer', '')),
                    'creation_date': str(metadata.get('/CreationDate', '')),
                    'modification_date': str(metadata.get('/ModDate', ''))
                },
                'text_length': len(text_content),
                'sensitive_data': sensitive_data,
                'target_mentions': text_content.lower().count(self.target.lower()),
                'analyzed_at': time.strftime('%Y-%m-%d %H:%M:%S')
            }
            
            return analysis
            
        except Exception as e:
            print_warning(f"Error analyzing PDF: {str(e)}")
            return None

    def _analyze_office_document(self, content, doc_info):
        """
        Analyze Office document content and metadata
        
        Args:
            content (bytes): Document content
            doc_info (dict): Document information
            
        Returns:
            dict: Office document analysis results
        """
        try:
            import io
            file_type = doc_info.get('file_type', '').lower()
            
            text_content = ''
            metadata = {}
            
            if file_type == 'docx' and DOCX_SUPPORT:
                # Analyze Word document
                doc_file = io.BytesIO(content)
                doc = DocxDocument(doc_file)
                
                # Extract text
                for paragraph in doc.paragraphs:
                    text_content += paragraph.text + '\n'
                
                # Extract metadata
                props = doc.core_properties
                metadata = {
                    'title': props.title or '',
                    'author': props.author or '',
                    'created': str(props.created) if props.created else '',
                    'modified': str(props.modified) if props.modified else '',
                    'subject': props.subject or '',
                    'keywords': props.keywords or ''
                }
            
            elif file_type in ['xls', 'xlsx'] and EXCEL_SUPPORT:
                # Analyze Excel document
                excel_file = io.BytesIO(content)
                workbook = load_workbook(excel_file)
                
                # Extract text from all sheets
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        for cell in row:
                            if cell:
                                text_content += str(cell) + ' '
                        text_content += '\n'
                
                # Extract metadata
                props = workbook.properties
                metadata = {
                    'title': props.title or '',
                    'creator': props.creator or '',
                    'created': str(props.created) if props.created else '',
                    'modified': str(props.modified) if props.modified else '',
                    'subject': props.subject or '',
                    'keywords': props.keywords or ''
                }
            
            # Analyze content
            sensitive_data = self._analyze_text_for_sensitive_data(text_content)
            
            analysis = {
                'url': doc_info.get('url', ''),
                'title': doc_info.get('title', ''),
                'file_type': file_type,
                'metadata': metadata,
                'text_length': len(text_content),
                'sensitive_data': sensitive_data,
                'target_mentions': text_content.lower().count(self.target.lower()),
                'analyzed_at': time.strftime('%Y-%m-%d %H:%M:%S')
            }
            
            return analysis
            
        except Exception as e:
            print_warning(f"Error analyzing Office document: {str(e)}")
            return None

    def _extract_document_metadata(self, content, file_type, doc_info):
        """
        Extract metadata from document
        
        Args:
            content (bytes): Document content
            file_type (str): File type
            doc_info (dict): Document information
            
        Returns:
            dict: Metadata information
        """
        # This would be called by the analysis functions above
        # Keeping as placeholder for future expansion
        return None

    def _scan_for_sensitive_content(self, content, doc_info):
        """
        Scan document content for sensitive information
        
        Args:
            content (bytes): Document content
            doc_info (dict): Document information
            
        Returns:
            list: List of sensitive content findings
        """
        sensitive_findings = []
        
        try:
            # Convert bytes to string for text analysis
            if isinstance(content, bytes):
                text_content = content.decode('utf-8', errors='ignore')
            else:
                text_content = str(content)
            
            # Scan for sensitive patterns
            for data_type, pattern in self.sensitive_patterns.items():
                matches = re.findall(pattern, text_content)
                
                if matches:
                    finding = {
                        'document_url': doc_info.get('url', ''),
                        'document_title': doc_info.get('title', ''),
                        'data_type': data_type,
                        'match_count': len(matches),
                        'sample_matches': matches[:3] if data_type != 'passwords' else ['***REDACTED***'] * min(3, len(matches)),
                        'discovered_at': time.strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    sensitive_findings.append(finding)
            
        except Exception as e:
            print_warning(f"Error scanning content for sensitive data: {str(e)}")
        
        return sensitive_findings

    def _analyze_text_for_sensitive_data(self, text_content):
        """
        Analyze text content for sensitive data patterns
        
        Args:
            text_content (str): Text to analyze
            
        Returns:
            list: List of sensitive data findings
        """
        sensitive_data = []
        
        for data_type, pattern in self.sensitive_patterns.items():
            matches = re.findall(pattern, text_content)
            
            if matches:
                sensitive_data.append({
                    'type': data_type,
                    'count': len(matches),
                    'samples': matches[:3] if data_type != 'passwords' else ['***REDACTED***'] * min(3, len(matches))
                })
        
        return sensitive_data

    def _extract_leaked_documents(self, html_content, source, pattern):
        """
        Extract leaked document information from search results
        
        Args:
            html_content (str): HTML content from search
            source (str): Source website
            pattern (str): Search pattern
            
        Returns:
            list: List of leaked document dictionaries
        """
        leaked_docs = []
        
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Extract search results
            results = soup.find_all('div', class_='g')  # Google results
            
            for result in results:
                title_elem = result.find('h3')
                title = title_elem.get_text().strip() if title_elem else 'No title'
                
                link_elem = result.find('a', href=True)
                url = link_elem['href'] if link_elem else ''
                
                if title and url and source in url:
                    leaked_docs.append({
                        'title': title,
                        'url': url
                    })
            
        except Exception as e:
            print_warning(f"Error extracting leaked documents from {source}: {str(e)}")
        
        return leaked_docs

    def _assess_leak_risk(self, doc):
        """
        Assess the risk level of a leaked document
        
        Args:
            doc (dict): Document information
            
        Returns:
            str: Risk level (high, medium, low)
        """
        title = doc.get('title', '').lower()
        
        high_risk_keywords = ['confidential', 'secret', 'private', 'internal', 'restricted']
        medium_risk_keywords = ['employee', 'staff', 'financial', 'contract', 'agreement']
        
        if any(keyword in title for keyword in high_risk_keywords):
            return 'high'
        elif any(keyword in title for keyword in medium_risk_keywords):
            return 'medium'
        else:
            return 'low'

    def run(self, output_format='json', output_dir='output'):
        """
        Run the document intelligence gathering
        
        Args:
            output_format (str): Output format (json, csv, txt)
            output_dir (str): Output directory
            
        Returns:
            dict: Results dictionary
        """
        print_info(f"Starting document intelligence gathering for: {self.target}")
        print_warning("DISCLAIMER: This module is for authorized security testing only!")
        
        try:
            # Discover public documents
            self.discover_public_documents()
            
            # Analyze discovered documents
            if self.results['discovered_documents']:
                self.analyze_discovered_documents()
            
            # Scan for leaked documents
            self.scan_for_leaked_documents()
            
            # Calculate statistics
            total_documents = len(self.results['discovered_documents'])
            total_pdf_analyzed = len(self.results['pdf_analysis'])
            total_office_analyzed = len(self.results['office_analysis'])
            total_sensitive_content = len(self.results['sensitive_content'])
            total_leaked = len(self.results['leaked_documents'])
            
            print_success(f"Document intelligence gathering completed!")
            print_info(f"Found {total_documents} documents")
            print_info(f"Analyzed {total_pdf_analyzed} PDF documents")
            print_info(f"Analyzed {total_office_analyzed} Office documents")
            print_info(f"Found {total_sensitive_content} sensitive content instances")
            print_info(f"Found {total_leaked} leaked documents")
            
            # Save results
            timestamp = time.strftime('%Y%m%d_%H%M%S')
            filename = f"doc_intel_{self.target}_{timestamp}"
            
            if 'json' in output_format:
                json_file = os.path.join(output_dir, f"{filename}.json")
                save_to_json(self.results, json_file)
                print_success(f"Results saved to: {json_file}")
            
            if 'csv' in output_format:
                csv_file = os.path.join(output_dir, f"{filename}.csv")
                self._save_to_csv(csv_file)
                print_success(f"Results saved to: {csv_file}")
            
            if 'txt' in output_format:
                txt_file = os.path.join(output_dir, f"{filename}.txt")
                self._save_to_txt(txt_file)
                print_success(f"Results saved to: {txt_file}")
            
            return self.results
            
        except KeyboardInterrupt:
            print_warning("Document intelligence gathering interrupted by user")
            return self.results
        except Exception as e:
            print_error(f"Error during document intelligence gathering: {str(e)}")
            return self.results

    def _save_to_csv(self, filename):
        """
        Save results to CSV format
        
        Args:
            filename (str): Output filename
        """
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Discovered documents
            writer.writerow(['Discovered Documents'])
            writer.writerow(['Title', 'URL', 'File Type', 'Is Sensitive', 'Source', 'Discovered At'])
            for doc in self.results['discovered_documents']:
                writer.writerow([
                    doc.get('title', ''),
                    doc.get('url', ''),
                    doc.get('file_type', ''),
                    doc.get('is_sensitive', ''),
                    doc.get('source', ''),
                    doc.get('discovered_at', '')
                ])
            
            writer.writerow([])
            
            # Leaked documents
            writer.writerow(['Leaked Documents'])
            writer.writerow(['Title', 'URL', 'Source', 'Risk Level', 'Discovered At'])
            for leak in self.results['leaked_documents']:
                writer.writerow([
                    leak.get('title', ''),
                    leak.get('url', ''),
                    leak.get('source', ''),
                    leak.get('risk_level', ''),
                    leak.get('discovered_at', '')
                ])

    def _save_to_txt(self, filename):
        """
        Save results to TXT format
        
        Args:
            filename (str): Output filename
        """
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f"Document Intelligence Gathering Results for: {self.target}\n")
            f.write(f"Generated: {self.results['timestamp']}\n")
            f.write("=" * 50 + "\n\n")
            
            f.write("DISCOVERED DOCUMENTS:\n")
            f.write("-" * 20 + "\n")
            for doc in self.results['discovered_documents']:
                f.write(f"Title: {doc.get('title', '')}\n")
                f.write(f"URL: {doc.get('url', '')}\n")
                f.write(f"File Type: {doc.get('file_type', '')}\n")
                f.write(f"Is Sensitive: {doc.get('is_sensitive', '')}\n")
                f.write(f"Source: {doc.get('source', '')}\n\n")
            
            f.write("LEAKED DOCUMENTS:\n")
            f.write("-" * 20 + "\n")
            for leak in self.results['leaked_documents']:
                f.write(f"Title: {leak.get('title', '')}\n")
                f.write(f"URL: {leak.get('url', '')}\n")
                f.write(f"Source: {leak.get('source', '')}\n")
                f.write(f"Risk Level: {leak.get('risk_level', '')}\n\n")
            
            f.write("SENSITIVE CONTENT:\n")
            f.write("-" * 20 + "\n")
            for content in self.results['sensitive_content']:
                f.write(f"Document: {content.get('document_title', '')}\n")
                f.write(f"Data Type: {content.get('data_type', '')}\n")
                f.write(f"Match Count: {content.get('match_count', '')}\n\n")


def main():
    """
    Main function for testing the document intelligence module
    """
    import argparse
    
    parser = argparse.ArgumentParser(description='NocturneRecon Document Intelligence Module')
    parser.add_argument('--target', '-t', required=True, help='Target domain or organization')
    parser.add_argument('--output', '-o', default='json', help='Output format (json, csv, txt)')
    parser.add_argument('--output-dir', '-d', default='output', help='Output directory')
    parser.add_argument('--threads', default=3, type=int, help='Number of threads')
    parser.add_argument('--delay', default=2.0, type=float, help='Delay between requests')
    parser.add_argument('--timeout', default=20, type=int, help='Request timeout')
    parser.add_argument('--max-file-size', default=10485760, type=int, help='Maximum file size to download (bytes)')
    
    args = parser.parse_args()
    
    config = {
        'threads': args.threads,
        'delay': args.delay,
        'timeout': args.timeout,
        'max_file_size': args.max_file_size
    }
    
    # Create output directory
    os.makedirs(args.output_dir, exist_ok=True)
    
    # Run document intelligence gathering
    gatherer = DocumentIntelligenceGatherer(args.target, config)
    results = gatherer.run(args.output, args.output_dir)
    
    print(f"\nDocument intelligence gathering completed for {args.target}")
    print(f"Results saved to {args.output_dir}/")


if __name__ == "__main__":
    main()
